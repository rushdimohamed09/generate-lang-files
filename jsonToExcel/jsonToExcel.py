# MIT License

# Copyright (c) 2023 Rushdi Mohamed

# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import os
import sys

# Get the parent directory of the current script (root folder)
root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Add the root folder to the Python module search path
sys.path.append(root_dir)
from utils.file_utils import read_json_file
import openpyxl

# Check if filename parameter was provided
# Load the JSON file
defaultFileName = 'sample.json'
data, filename = read_json_file(defaultFileName)

# Create a new workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active

# Set the header row
ws.append(['keyName', 'text'])
ws.append([])

# Add the data
for i, key in enumerate(data.keys(), start=3):
    ws[f"A{i}"] = key
    ws[f"B{i}"] = data[key]

# Save the workbook
languageDataFile = 'jsonToExcel/languageData.xlsx'
wb.save(languageDataFile)
print(f"{languageDataFile} file generated successfully based on {filename}!")
